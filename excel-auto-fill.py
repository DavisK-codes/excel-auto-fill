import os
import shutil
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
import re

print('To exit the program type "exit"')

input_folder = "Input"
output_folder = "Output"
template_folder = "Template"
possible_extensions = ['.xls', '.xlsx']

while True:
    input_filename = input('Enter the file name:')
    
    if input_filename == 'exit':
        quit()
    for ext in possible_extensions:
        temp_file = os.path.join(input_folder, input_filename + ext)
        if os.path.exists(temp_file):
            input_file = temp_file
            print(f'Trying to read: {input_file}')
            break
    else:
        print('File not found!')
        continue
    try:
        file_extension = os.path.splitext(input_file)[1].lower()
        if file_extension == '.xlsx':
            df = pd.read_excel(input_file, engine="openpyxl", header=1)
            break
        elif file_extension == '.xls':
            df = pd.read_excel(input_file, engine="xlrd")
            converted_file = input_file.replace(".xls", ".xlsx")  
            df.to_excel(converted_file, index=False, engine="openpyxl")
            input_file = converted_file
            break
        else:
            print(f"Unsupported file format: {file_extension}")
            continue

    except Exception as error:
        print('Error occurred while processing the file!')
        print(error)
        continue

#                                 RETRIEVING DATA FROM THE INPUT FILE


# Load the workbook
wb = load_workbook(input_file)
ws = wb.active

extracted_data = []

# Iterate over rows dynamically
for row in ws.iter_rows(min_row=1, max_col=14):  # Adjust max_col if needed
    row_data = [cell.value for cell in row]
    
    # Check if the row contains any meaningful value
    # Only append rows that have at least one non-None/empty value
    if any(cell is not None and cell != "" for cell in row_data):
        extracted_data.append(row_data)
    else:
        continue

# For data that appears only once in the sheet (single-row values, no looping needed)
data = {
    'subject': None,
    'additional_information': None
}

item_list = []

# This pattern is used to match rows that have similar data across all of them, 
# specifically, in our input file, this corresponds to row IDs in Excel, like (A40).
pattern = r'^=ROW\(A\d+\)\s*-\s*\d+$'

for row in extracted_data:
    if row is not None and len(row) > 0:
        if 'Subject: ' in row:
            data['subject'] = row[10] if row[10] else data['subject']
        
        #if the row matches our pattern, extract item data
        if row[0] and isinstance(row[0], str) and re.match(pattern, row[0]):
            try:
                match = re.match(r'^=ROW\([A-Z]+(\d+)\)\s*-\s*(\d+)$', row[0])
                row_number = int(match.group(1))  # Extract the number from ROW(A40)
                subtract_value = int(match.group(2))  # Extract the number being subtracted (39)
                item_number = row_number - subtract_value
                
                # Create a dictionary for item data, using appropriate values from the row
                item_data = {
                'item_number' : item_number,
                'description' : row[1].strip() if row[1] else None,
                'code_maker_ref' : row[4] if row[4] else None,
                'part_nr_maker' : row[5].strip() if row[5] else None,
                'part_type' : row[6].strip() if row[6] else None,
                'serial_nr' : row[7] if row[7] else None,
                'qty' : float(row[8]) if row[8] else None,
                'unit' : row[9].strip() if row[9] else None
                }
                
                
                item_list.append(item_data)

            except ValueError:
                continue
try:
    raw_additional_information = extracted_data[12][0]
    # The data in extracted_data[12][0] may contain unwanted encoded patterns like '_x000D_'.
    # The regular expression replaces these patterns with a space, and strip() removes any leading or trailing spaces.
    # This ensures we clean the data before storing it in data['additional_information'].
    data['additional_information'] = re.sub(r'_x[0-9A-Fa-f]{4}_', ' ', raw_additional_information).strip()
except:
    data['additional_information'] = None


# Ensure the output folder exists
os.makedirs(output_folder, exist_ok=True)

# Define the paths
template_file = os.path.join(template_folder, 'test.xlsx')
output_file = os.path.join(output_folder, 'Qutation.xlsx')

# Copy template to output file
shutil.copy(template_file, output_file)

# Load the new file (copied from the template)
wb_output = load_workbook(output_file)
ws_output = wb_output.active
print(f"Template copied successfully to: {output_file}")


#                                     STORING DATA IN THE OUTPUT FILE


for row in ws_output.iter_rows(min_row=1, max_col=14):
    row_data = [cell.value for cell in row]
   

# Step 1: Define the start and end for the first and last row where items are to be inserted
row_num = 10
last_item_row = row_num + len(item_list) - 1  # Last item row
insert_amount = len(item_list) # Number of rows to insert based on item list

# Step 2: Define the location in our case for additional information and delivery notes boxes
# in the output file. They are placed two rows below the last item.
boxes_start = row_num + 2
box_content = {}

# Step 3: Identify and unmerge merged cells that are below the last item.
merged_ranges_to_move = []

# Used to identify the locations of merged cells (e.g., additional information box and delivery notes box)
# in the output file. These merged cells could interfere with the insertion of new rows, so we need to handle them.
for merged_range in list(ws_output.merged_cells.ranges):
    if merged_range.min_row >= boxes_start:
        top_left_cell = ws_output.cell(merged_range.min_row, merged_range.min_col)
        box_content[(merged_range.min_row, merged_range.min_col)] = top_left_cell.value
        
        # Add the merged range information to the list to move
        merged_ranges_to_move.append((
            merged_range.coord, 
            merged_range.min_row, merged_range.max_row, 
            merged_range.min_col, merged_range.max_col
        ))
        
        # Unmerge the cells so we can reposition them correctly
        ws_output.unmerge_cells(merged_range.coord)

# Step 4: Insert Item rows and rows below last item for additional and delivery note boxes

for coord, old_min_row, old_max_row, old_min_col, old_max_col in merged_ranges_to_move:
    new_min_row = old_min_row + insert_amount
    new_max_row = old_max_row + insert_amount
    new_min_col = get_column_letter(old_min_col)
    new_max_col = get_column_letter(old_max_col)

    new_coord = f"{new_min_col}{new_min_row}:{new_max_col}{new_max_row}"
    ws_output.merge_cells(new_coord)

ws_output.insert_rows(boxes_start, amount=insert_amount)

# Step 4: Load the data
ws_output['A5'] = data['subject']  

# Insert Item Data
for item in item_list:
    ws_output[f'A{row_num}'] = item['item_number']  
    ws_output[f'B{row_num}'] = item['description'] 
    ws_output[f'E{row_num}'] = item['code_maker_ref']
    ws_output[f'F{row_num}'] = item['part_nr_maker']
    ws_output[f'G{row_num}'] = item['part_type']
    ws_output[f'H{row_num}'] = item['serial_nr']
    ws_output[f'I{row_num}'] = item['qty']
    ws_output[f'J{row_num}'] = item['unit']

    row_num += 1 

# Calculating how big the additional and delivery note box should be
# so that all the data inside it is visible
additional_info_lines = data['additional_information'].split('\n')  
num_lines = len(additional_info_lines)
# Calculate row height based on the number of lines
line_height = 14.5
calculated_height = num_lines * line_height

# Adjust the row height dynamically
# Here, we'll set the height for the row where additional_information is placed
additional_info_row = row_num + 5
ws_output.row_dimensions[additional_info_row].height = calculated_height

# Ensure Additional Information is Correctly Placed
ws_output[f'B{additional_info_row}'] = data['additional_information']

wb_output.save(output_file)
print(f"Finished data inserted and saved in: {output_file}")


#                                      FORMATTING + STYLE

# Load the workbook
wb_output = openpyxl.load_workbook(output_file)
ws_output = wb_output.active 

# Define font, alignment, and border style
tahoma_font  = Font(name='Tahoma', size=11)
border_style = Border(
    left=Side(style='thin', color="E4E0D8"),
    right=Side(style='thin', color="E4E0D8"),
    top=Side(style='thin', color="E4E0D8"),
    bottom=Side(style='thin', color="E4E0D8")
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
middle_left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

starting_row = 10
# Apply styling and formatting to the item rows
for idx, item in enumerate(item_list):
    # Calculate the current row for this item
    row = starting_row + idx
    # Ensure Excel automatically adjusts row height
    ws_output.row_dimensions[row].height = None

    # Item Number (A column)
    cell = ws_output[f'A{row}']
    cell.font = tahoma_font
    cell.alignment = center_align
    cell.border = border_style
    cell.value = item['item_number']

    # Description (B, C, D columns)
    cell = ws_output[f'B{row}']
    cell.font = tahoma_font
    cell.alignment = left_align
    cell.border = border_style
    cell.value = item['description'].upper() if isinstance(item['description'], str) else item['description']  # Only upper if it's a string
    ws_output.merge_cells(f'B{row}:D{row}')  # Merge B to D for description

    # Code Maker Ref (E column)
    cell = ws_output[f'E{row}']
    cell.font = tahoma_font
    cell.alignment = middle_left_align
    cell.border = border_style
    cell.value = item['code_maker_ref'].upper() if isinstance(item['code_maker_ref'], str) else item['code_maker_ref']
    
    # Part Nr Maker (F column)
    cell = ws_output[f'F{row}']
    cell.font = tahoma_font
    cell.alignment = middle_left_align
    cell.border = border_style
    cell.value = item['part_nr_maker'].upper() if isinstance(item['part_nr_maker'], str) else item['part_nr_maker']
    
    # Part Type (G column)
    cell = ws_output[f'G{row}']
    cell.font = tahoma_font
    cell.alignment = middle_left_align
    cell.border = border_style
    cell.value = item['part_type'].upper() if isinstance(item['part_type'], str) else item['part_type']
    
    # Serial Nr (H column)
    cell = ws_output[f'H{row}']
    cell.font = tahoma_font
    cell.alignment = middle_left_align
    cell.border = border_style
    cell.value = item['serial_nr'].upper() if isinstance(item['serial_nr'], str) else item['serial_nr'] 

    # Quantity (I column)
    cell = ws_output[f'I{row}']
    cell.font = tahoma_font
    cell.alignment = center_align
    cell.border = border_style
    cell.value = item['qty']

    # Collumns J - N
    for col in range (10, 15):
        cell = ws_output.cell(row=row, column=col)
        cell.font = tahoma_font
        cell.alignment = center_align
        cell.border = border_style
    
    # Column M - Percentage format
    cell = ws_output[f'M{row}']
    cell.number_format = '0.00%'  # Format as percentage (e.g., 50.00%)
    
    # Column K and N - Number format
    cell = ws_output[f'K{row}']
    cell.number_format = '0.00'

    cell = ws_output[f'N{row}']
    cell.number_format = '0.00' 


# Step 6: Save the workbook after applying formatting
wb_output.save(output_file)
print(f"Finished formatting and saving: {output_file}")

#                                          PROTECTING WORKBOOK


sheet_name = "Quotation Form"

if sheet_name in wb_output.sheetnames:
    ws = wb_output[sheet_name]

    # Step 1: Unlock all cells first
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = Protection(locked=False)

    # Step 2: Locking the cells
    
    rows_to_lock = [3, 5, 7, 9]
    protected_columns_A_to_N = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]
    for row_num in rows_to_lock:
        for col in protected_columns_A_to_N:
            cell_ref = f"{col}{row_num}"
            ws[cell_ref].protection = Protection(locked=True)  # Lock the cell

    # Columns within item rows which needs protection
    protected_columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]
    # After protecting cells, Excel might remove any existing cell fill. We redefine the white fill 
    # to ensure that the protected cells still have the correct background color.
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for idx, item in enumerate(item_list):
        row_num = starting_row + idx  # Calculate the row dynamically

        for col in protected_columns:
            cell_ref = f"{col}{row_num}"
            ws[cell_ref].protection = Protection(locked=True) 
            ws[cell_ref].fill = white_fill  # Apply white fill to the protected cells
    
    additional_info_rows = [row_num + 4, row_num + 6]
    additional_info_columns = ["B", "C", "D", "E", "F"]
    
    for row_num in additional_info_rows:
        for col in additional_info_columns:
            cell_ref = f"{col}{row_num}"
            ws[cell_ref].protection = Protection(locked=True)
    
    # Filling additional rows around Additional and Delivery note boxes
    row_num = row_num - 6
    additional_rows_to_fill = 3
    non_protected_rows = range(row_num + 1, row_num + 1 + additional_rows_to_fill)

    for row_num in non_protected_rows:
        for col in ws.columns:  
            cell = ws.cell(row=row_num, column=col[0].column) 
            cell.fill = white_fill 
    
    extra_row_to_fill = last_item_row + 5
    for col in ws.columns:  
        cell = ws.cell(row=extra_row_to_fill, column=col[0].column)  
        cell.fill = white_fill 
    

   # Step 6: Enable sheet protection
    ws.protection.sheet = True
    ws.protection.password = "cayenne" 
 

wb_output.save(output_file)

# After saving the workbook, prompt the program to open the output file and then close the program

if os.name == "nt":  # Windows
    os.startfile(output_file)
elif sys.platform == "darwin":  # macOS
    subprocess.run(["open", output_file])

    # Close Terminal only on macOS
    os.system("osascript -e 'tell application \"Terminal\" to close front window'")

quit() 